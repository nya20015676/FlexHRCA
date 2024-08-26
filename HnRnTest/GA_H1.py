import sys

sys.path.append('..')  
sys.path.append('../opTimeData')  
from hrcStation import Human, Robot, Productt, HrcStation
from fuzzyNumber import Triangular, SVTNN
from scipy.stats import norm, nbinom
from openpyxl import Workbook
import pickle
import os
import send2trash
import itertools
import copy
from time_data import TimeData
from Time import TIME_TABLE


human1 = Human("human1", capability = 3)
human2 = Human("human2", capability = 6)
human3 = Human("human3", capability = 6)
human4 = Human("human4", capability = 6)
human5 = Human("human5", capability = 6)
robot1 = Robot("robot1", capability = 3)
robot2 = Robot("robot2", capability = 5)
robot3 = Robot("robot3", capability = 8)
robot4 = Robot("robot4", capability = 8)
robot5 = Robot("robot5", capability = 8)
station_static = HrcStation()
station_fuzzy = HrcStation()

#Product
n_job_op = {0:3, 1:3, 2:2, 3:2, 4:1}
after_require = [ # 0~2 Battery1 installation
                    [], [0], [0],
                  # 3~5 Battery2 installation
                    [], [3], [3],
                  # 6~7 Controller1 installation
                    [], [6], 
                  # 8~9 Controller2 installation
                    [], [8],  ]
after_require.append(list(range(10)))


def transTime_SVTNN2static(Time):
    new_Time = copy.deepcopy(Time)
    for lst in new_Time:
      for i in range(len(lst)):
          if lst[i] != 9999:
              lst[i] = lst[i].mean
    return new_Time
def transTime_SVTNN2fuzzy(Time):
    new_Time = copy.deepcopy(Time)
    for lst in new_Time:
      for i in range(len(lst)):
          if lst[i] != 9999:
            lst[i] = Triangular(lst[i].l_std, lst[i].mean, lst[i].r_std)
    return new_Time
def setWeight_SVTNN(Time, weight):
    new_Time = copy.deepcopy(Time)
    for lst in new_Time:
      for i in range(len(lst)):
          if lst[i] != 9999:
            lst[i].set_weight(weight, 1-weight, 1-weight)
    return new_Time



def excel_writing(filename, data):

    wb = Workbook()
    ws = wb.active

    # 写入列标题
    column_titles = ['Iteration','Low', 'Optimal Fitness', 'Up','Low', 'Best Fitness', 'Up', 'Mean Fitness', 'Offspring Number', 'Time', 'Optimal Size']
    for col, title in enumerate(column_titles, start=1):
        ws.cell(row=1, column=col, value=title)

    # 写入具体数值
    for row, item in enumerate(data, start=2):
        for col, value in enumerate(item, start=1):
            ws.cell(row=row, column=col, value=value)

    # 保存工作簿
    if os.path.exists(filename):
        send2trash.send2trash(filename)  
    wb.save(filename)

def printable(record_list):
    # record_list: [[iteration, low, optimal_fitness, up, low, best_fitness, up, mean_fitness, pop_num, time, optimal_size], ...]
    for record in record_list:
        optimal_fitness = record[1]
        if type(optimal_fitness) != SVTNN:
            best_fitness = record[2]
            record.insert(1, optimal_fitness)
            record.insert(3, optimal_fitness)
            record.insert(4, best_fitness)
            record.insert(6, best_fitness)
            record[2] = optimal_fitness
            record[5] = best_fitness
            for i in range(len(record)):
                record[i] = round(record[i])
        else:
            best_fitness = record[2]
            record.insert(1, optimal_fitness.low)
            record.insert(3, optimal_fitness.up)
            record.insert(4, best_fitness.low)
            record.insert(6, best_fitness.up)
            record[2] = optimal_fitness.mean
            record[5] = best_fitness.mean
            for i in range(len(record)):
                record[i] = round(record[i])

# station.gantt(OS_COS_MS_TS_START_END_AFTER)

# p_cross_machine 变异了MS就不会变异OS了
pop_size_list = [100, 200,300]
p_crossover_list = [0.7,0.8,0.9]
p_cross_machine_list = [0.5]
p_mutation_list = [0.1,0.2,0.3]
p_mutation_machine_list = [0.5]
max_iteration_list = [1000]
tSize_list = [3,4,5]

path_list = ["/home/amtc/Desktop/optimal/HnRnTest/H1/H1R1/",
             "/home/amtc/Desktop/optimal/HnRnTest/H1/H1R2/",
             "/home/amtc/Desktop/optimal/HnRnTest/H1/H1R3/",
             "/home/amtc/Desktop/optimal/HnRnTest/H1/H1R4/",
             "/home/amtc/Desktop/optimal/HnRnTest/H1/H1R5/",]

Time_list = [TIME_TABLE.Time_H1R1, TIME_TABLE.Time_H1R2, TIME_TABLE.Time_H1R3, TIME_TABLE.Time_H1R4, TIME_TABLE.Time_H1R5]
robots_list = [[robot1], [robot1, robot2], [robot1, robot2, robot3], [robot1, robot2, robot3, robot4], [robot1, robot2, robot3, robot4, robot5]]

for i in [4,5]:
    for j, Time in enumerate(Time_list):
        product1 = Productt("Battery pack")
        product2 = Productt("Battery pack")
        product3 = Productt("Battery pack")
        product1.set_jobs(n_job_op, after_require)
        product2.set_jobs(n_job_op, after_require)
        product3.set_jobs(n_job_op, after_require)
        station_SVTNN = HrcStation()
        weight = 0.25
        Time = setWeight_SVTNN(Time, weight)
        station_SVTNN.set_operator_product_Time([human1], robots_list[j], [product1, product2, product3], [Time, Time, Time])
        
        if __name__ == '__main__':
            optimal_fitness, mem_SVTNN_CHS_list, mem_SVTNN_CHS_fitness_list,  record_list = station_SVTNN.GA(pop_size=300, p_crossover=0.9, p_cross_machine=0.67, p_mutation=0.1, p_mutation_machine=0.83, max_iteration=1000, tSize=4)
            # optimal_fitness, mem_SVTNN_CHS_list, mem_SVTNN_CHS_fitness_list,  record_list = station_SVTNN.GA_initialed_pop(mem_static_CHS_list, pop_size=300, p_crossover=0.9, p_cross_machine=0.67, p_mutation=0.1, p_mutation_machine=0.83, max_iteration=300, tSize=4)
            
            printable(record_list)
            excel_name = path_list[j] + "{}-{}.xlsx".format(i, optimal_fitness.mean if isinstance(optimal_fitness, SVTNN) else optimal_fitness)
            data_name = path_list[j] + "{}-{}.pickle".format(i, optimal_fitness.mean if isinstance(optimal_fitness, SVTNN) else optimal_fitness)
            excel_writing(excel_name, record_list)
            with open(data_name, "wb") as file:
                pickle.dump(mem_SVTNN_CHS_list, file)