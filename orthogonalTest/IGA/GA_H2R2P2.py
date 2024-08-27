import sys

sys.path.append('..')  
sys.path.append('../opTimeData')  
from hrcStation import Human, Robot, Productt, HrcStation
from fuzzyNumber import Triangular, SVTNN
from scipy.stats import norm, nbinom
from openpyxl import Workbook
import pickle
import os
import send2trash
import itertools
import copy
from time_data import TimeData


human1 = Human("human1", capability = 3)
human2 = Human("human2", capability = 6)
human3 = Human("human3", capability = 6)
human4 = Human("human4", capability = 6)
human5 = Human("human5", capability = 6)
robot1 = Robot("robot1", capability = 3)
robot2 = Robot("robot2", capability = 5)
robot3 = Robot("robot3", capability = 8)
robot4 = Robot("robot4", capability = 8)
robot5 = Robot("robot5", capability = 8)
station_static = HrcStation()
station_fuzzy = HrcStation()
station_SVTNN = HrcStation()

#Product
n_job_op = {0:3, 1:3, 2:2, 3:2, 4:1}
after_require = [ # 0~2 Battery1 installation
                    [], [0], [0],
                  # 3~5 Battery2 installation
                    [], [3], [3],
                  # 6~7 Controller1 installation
                    [], [6], 
                  # 8~9 Controller2 installation
                    [], [8],  ]
after_require.append(list(range(10)))
product1 = Productt("Battery pack")
product2 = Productt("Battery pack")
product3 = Productt("Battery pack")
product4 = Productt("Battery pack")
product5 = Productt("Battery pack")
product6 = Productt("Battery pack")
product7 = Productt("Battery pack")
product8 = Productt("Battery pack")
product9 = Productt("Battery pack")
product1.set_jobs(n_job_op, after_require)
product2.set_jobs(n_job_op, after_require)
product3.set_jobs(n_job_op, after_require)
product4.set_jobs(n_job_op, after_require)
product5.set_jobs(n_job_op, after_require)
product6.set_jobs(n_job_op, after_require)
product7.set_jobs(n_job_op, after_require)
product8.set_jobs(n_job_op, after_require)
product9.set_jobs(n_job_op, after_require)

# Time:
    #     H1 R1 R2 H1&R1 H1&R2 
    # O11 t  t  9999
    # O12
    # ...
    # 9999代表不能选择

Time_H2R2 = [
    # Battery1 installation
    # Pick and place battery1
    [TimeData.time_O11_H1, TimeData.time_O11_H2, TimeData.time_O11_R1, TimeData.time_O11_R2, 9999, 9999, 9999, 9999],
    # Tighten 4 screws of left side
    [TimeData.time_O12_H1, TimeData.time_O12_H2, TimeData.time_O12_R1, TimeData.time_O12_R2, 9999, 9999, 9999, 9999],
    # Tighten 4 screws of right side
    [TimeData.time_O13_H1, TimeData.time_O13_H2, TimeData.time_O13_R1, TimeData.time_O13_R2, 9999, 9999, 9999, 9999],

    # Battery2 installation
    # Pick and place battery2
    [TimeData.time_O21_H1, TimeData.time_O21_H2, TimeData.time_O21_R1, TimeData.time_O21_R2, 9999, 9999, 9999, 9999],
    # Tighten 4 screws of left side
    [TimeData.time_O22_H1, TimeData.time_O22_H2, TimeData.time_O22_R1, TimeData.time_O22_R2, 9999, 9999, 9999, 9999],
    # Tighten 4 screws of right side
    [TimeData.time_O23_H1, TimeData.time_O23_H2, TimeData.time_O23_R1, TimeData.time_O23_R2, 9999, 9999, 9999, 9999],

    # Controller1 installation
    # Pick, prepare and place controller1
    [TimeData.time_O31_H1, TimeData.time_O31_H2, TimeData.time_O31_R1, TimeData.time_O31_R2, 9999, 9999, 9999, 9999],
    # Tighten 4 screws
    [TimeData.time_O32_H1, TimeData.time_O32_H2, TimeData.time_O32_R1, TimeData.time_O32_R2, 9999, 9999, 9999, 9999],
    
    # Controller2 installation
    # Pick, prepare and place controller2
    [TimeData.time_O41_H1, TimeData.time_O41_H2, TimeData.time_O41_R1, TimeData.time_O41_R2, 9999, 9999, 9999, 9999],
    # Tighten 4 screws
    [TimeData.time_O42_H1, TimeData.time_O42_H2, TimeData.time_O42_R1, TimeData.time_O42_R2, 9999, 9999, 9999, 9999],

    # Cables installation
    [9999, 9999, 9999, 9999, TimeData.time_O51_H1, 9999, TimeData.time_O51_H2, 9999],
]

def transTime_SVTNN2static(Time):
    new_Time = copy.deepcopy(Time)
    for lst in new_Time:
      for i in range(len(lst)):
          if lst[i] != 9999:
              lst[i] = lst[i].mean
    return new_Time
def transTime_SVTNN2fuzzy(Time):
    new_Time = copy.deepcopy(Time)
    for lst in new_Time:
      for i in range(len(lst)):
          if lst[i] != 9999:
            lst[i] = Triangular(lst[i].l_std, lst[i].mean, lst[i].r_std)
    return new_Time

station_static.set_operator_product_Time([human1, human2], [robot1, robot2], [product1, product2], [transTime_SVTNN2static(Time_H2R2), transTime_SVTNN2static(Time_H2R2)])
station_fuzzy.set_operator_product_Time([human1, human2], [robot1, robot2], [product4, product5], [transTime_SVTNN2fuzzy(Time_H2R2), transTime_SVTNN2fuzzy(Time_H2R2)])
station_SVTNN.set_operator_product_Time([human1, human2], [robot1, robot2], [product7, product8], [Time_H2R2, Time_H2R2])

def excel_writing(filename, data):

    wb = Workbook()
    ws = wb.active

    # 写入列标题
    column_titles = ['Iteration','Low', 'Optimal Fitness', 'Up','Low', 'Best Fitness', 'Up', 'Mean Fitness', 'Offspring Number', 'Time', 'Optimal Size']
    for col, title in enumerate(column_titles, start=1):
        ws.cell(row=1, column=col, value=title)

    # 写入具体数值
    for row, item in enumerate(data, start=2):
        for col, value in enumerate(item, start=1):
            ws.cell(row=row, column=col, value=value)

    # 保存工作簿
    if os.path.exists(filename):
        send2trash.send2trash(filename)  
    wb.save(filename)

def printable(record_list):
    # record_list: [[iteration, low, optimal_fitness, up, low, best_fitness, up, mean_fitness, pop_num, time, optimal_size], ...]
    for record in record_list:
        optimal_fitness = record[1]
        if type(optimal_fitness) != SVTNN:
            best_fitness = record[2]
            record.insert(1, optimal_fitness)
            record.insert(3, optimal_fitness)
            record.insert(4, best_fitness)
            record.insert(6, best_fitness)
            record[2] = optimal_fitness
            record[5] = best_fitness
            for i in range(len(record)):
                record[i] = round(record[i])
        else:
            best_fitness = record[2]
            record.insert(1, optimal_fitness.low)
            record.insert(3, optimal_fitness.up)
            record.insert(4, best_fitness.low)
            record.insert(6, best_fitness.up)
            record[2] = optimal_fitness.mean
            record[5] = best_fitness.mean
            for i in range(len(record)):
                record[i] = round(record[i])

# station.gantt(OS_COS_MS_TS_START_END_AFTER)

# p_cross_machine 变异了MS就不会变异OS了
pop_size_list = [100, 200,300]
p_crossover_list = [0.7,0.8,0.9]
p_cross_machine_list = [0.5]
p_mutation_list = [0.1,0.2,0.3]
p_mutation_machine_list = [0.5]
max_iteration_list = [1000]
tSize_list = [3,4,5]



# 生成所有参数的组合
combinations_old = [
    [100, 0.7, 0.1, 3, 0.3, 0.7, 1000],
    [100, 0.8, 0.2, 4, 0.3, 0.7, 1000],
    [100, 0.9, 0.3, 5, 0.3, 0.7, 1000],
    [200, 0.7, 0.2, 5, 0.3, 0.7, 1000],
    [200, 0.8, 0.3, 3, 0.3, 0.7, 1000],
    [200, 0.9, 0.1, 4, 0.3, 0.7, 1000],
    [300, 0.7, 0.3, 4, 0.3, 0.7, 1000],
    [300, 0.8, 0.1, 5, 0.3, 0.7, 1000],
    [300, 0.9, 0.2, 3, 0.3, 0.7, 1000],
]

combinations = [
    [200, 0.5, 0.1, 4, 0.16, 0.16, 1000],
    [200, 0.6, 0.2, 5, 0.33, 0.33, 1000],
    [200, 0.7, 0.3, 6, 0.50, 0.50, 1000],
    [200, 0.8, 0.4, 7, 0.67, 0.67, 1000],
    [200, 0.9, 0.5, 8, 0.83, 0.83, 1000],
    [250, 0.5, 0.3, 8, 0.33, 0.67, 1000],
    [250, 0.6, 0.4, 4, 0.50, 0.83, 1000],
    [250, 0.7, 0.5, 5, 0.67, 0.16, 1000],
    [250, 0.8, 0.1, 6, 0.83, 0.33, 1000],
    [250, 0.9, 0.2, 7, 0.16, 0.50, 1000],
    [300, 0.5, 0.5, 7, 0.50, 0.33, 1000],
    [300, 0.6, 0.1, 8, 0.67, 0.50, 1000],
    [300, 0.7, 0.2, 4, 0.83, 0.67, 1000],
    [300, 0.8, 0.3, 5, 0.16, 0.83, 1000],
    [300, 0.9, 0.4, 6, 0.33, 0.16, 1000],
    [350, 0.5, 0.2, 6, 0.67, 0.83, 1000],
    [350, 0.6, 0.3, 7, 0.83, 0.16, 1000],
    [350, 0.7, 0.4, 8, 0.16, 0.33, 1000],
    [350, 0.8, 0.5, 4, 0.33, 0.50, 1000],
    [350, 0.9, 0.1, 5, 0.50, 0.67, 1000],
    [400, 0.5, 0.4, 5, 0.83, 0.50, 1000],
    [400, 0.6, 0.5, 6, 0.16, 0.67, 1000],
    [400, 0.7, 0.1, 7, 0.33, 0.83, 1000],
    [400, 0.8, 0.2, 8, 0.50, 0.16, 1000],
    [400, 0.9, 0.3, 4, 0.67, 0.33, 1000],
]

path = "/home/amtc/Desktop/optimal/orthogonalTest/results/H2R2P2/" # 文件保存路径

for i in [1,2,3,4,5]:
    for pop_size, p_crossover, p_mutation, tSize, p_cross_machine, p_mutation_machine, max_iteration in combinations:
        if __name__ == '__main__':
            mem_static_CHS_list = []

            optimal_fitness, mem_static_CHS_list, mem_static_CHS_fitness_list, record_list = station_static.GA( pop_size, p_crossover, p_cross_machine, p_mutation, p_mutation_machine, max_iteration, tSize)
            # mem_fuzzy_CHS_list = []
            # while len(mem_fuzzy_CHS_list) < 100:
            #   optimal_fitness, mem_fuzzy_CHS_list, mem_fuzzy_CHS_fitness_list, record_list = station_fuzzy.GA_initialed_pop(mem_static_CHS_list, pop_size, p_crossover, p_cross_machine, p_mutation, p_mutation_machine, max_iteration)
            # mem_SVTNN_CHS_list = []
            
            # optimal_fitness, mem_SVTNN_CHS_list, mem_SVTNN_CHS_fitness_list,  record_list = station_SVTNN.GA_initialed_pop(mem_fuzzy_CHS_list, pop_size, p_crossover, p_cross_machine, p_mutation, p_mutation_machine, max_iteration)
            
            printable(record_list)
            excel_name = path +"{}/".format(i) + "{}-popSize={}_pCrossover={}_pCrossMS={}_pMutate={}_pMutateMS={}_maxIteration={}.xlsx".format(optimal_fitness.mean if isinstance(optimal_fitness, SVTNN) else optimal_fitness, pop_size, p_crossover, p_cross_machine, p_mutation, p_mutation_machine, max_iteration)
            data_name = path +"{}/".format(i) + "{}-popSize={}_pCrossover={}_pCrossMS={}_pMutate={}_pMutateMS={}_maxIteration={}.pickle".format(optimal_fitness.mean if isinstance(optimal_fitness, SVTNN) else optimal_fitness, pop_size, p_crossover, p_cross_machine, p_mutation, p_mutation_machine, max_iteration)
            excel_writing(excel_name, record_list)
            with open(data_name, "wb") as file:
                pickle.dump(mem_static_CHS_list, file)