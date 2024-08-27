import sys
# 添加上一级目录到Python的模块搜索路径中
sys.path.append('..')  # 将上一级目录添加到Python的模块搜索路径中
sys.path.append('../opTimeData')  # 将上一级目录添加到Python的模块搜索路径中
from hrcStation import Human, Robot, Productt, HrcStation
from fuzzyNumber import Triangular, SVTNN
from scipy.stats import norm, nbinom
from openpyxl import Workbook
import pickle
import os
import send2trash
import itertools
import copy
from time_data import TimeData,SVTN_682_954,SVTN_682_997,SVTN_954_997
import random

# capability越大opTime越慢
human1 = Human("human1", capability = 3)
human2 = Human("human2", capability = 6)
human3 = Human("human3", capability = 6)
human4 = Human("human4", capability = 6)
human5 = Human("human5", capability = 6)
robot1 = Robot("robot1", capability = 3)
robot2 = Robot("robot2", capability = 5)
robot3 = Robot("robot3", capability = 8)
robot4 = Robot("robot4", capability = 8)
robot5 = Robot("robot5", capability = 8)
station_static = HrcStation()
station_fuzzy = HrcStation()

#Product
n_job_op = {0:3, 1:3, 2:2, 3:2, 4:1}
after_require = [ # 0~2 Battery1 installation
                    [], [0], [0],
                  # 3~5 Battery2 installation
                    [], [3], [3],
                  # 6~7 Controller1 installation
                    [], [6], 
                  # 8~9 Controller2 installation
                    [], [8],  ]
after_require.append(list(range(10)))


# Time:
    #     H1 R1 R2 H1&R1 H1&R2 
    # O11 t  t  9999
    # O12
    # ...
    # 9999代表不能选择

Time_H2R3_682_954 = [
    # Battery1 installation
    # Pick and place battery1
    [SVTN_682_954.time_O11_H1, SVTN_682_954.time_O11_H2, SVTN_682_954.time_O11_R1, SVTN_682_954.time_O11_R2, SVTN_682_954.time_O11_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    # Tighten 4 screws of left side
    [SVTN_682_954.time_O12_H1, SVTN_682_954.time_O12_H2, SVTN_682_954.time_O12_R1, SVTN_682_954.time_O12_R2, SVTN_682_954.time_O12_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    # Tighten 4 screws of right side
    [SVTN_682_954.time_O13_H1, SVTN_682_954.time_O13_H2, SVTN_682_954.time_O13_R1, SVTN_682_954.time_O13_R2, SVTN_682_954.time_O13_R3, 9999, 9999, 9999, 9999, 9999, 9999],

    # Battery2 installation
    # Pick and place battery2
    [SVTN_682_954.time_O21_H1, SVTN_682_954.time_O21_H2, SVTN_682_954.time_O21_R1, SVTN_682_954.time_O21_R2, SVTN_682_954.time_O21_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    # Tighten 4 screws of left side
    [SVTN_682_954.time_O22_H1, SVTN_682_954.time_O22_H2, SVTN_682_954.time_O22_R1, SVTN_682_954.time_O22_R2, SVTN_682_954.time_O22_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    # Tighten 4 screws of right side
    [SVTN_682_954.time_O23_H1, SVTN_682_954.time_O23_H2, SVTN_682_954.time_O23_R1, SVTN_682_954.time_O23_R2, SVTN_682_954.time_O23_R3, 9999, 9999, 9999, 9999, 9999, 9999],

    # Controller1 installation
    # Pick, prepare and place controller1
    [SVTN_682_954.time_O31_H1, SVTN_682_954.time_O31_H2, SVTN_682_954.time_O31_R1, SVTN_682_954.time_O31_R2, SVTN_682_954.time_O31_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    # Tighten 4 screws
    [SVTN_682_954.time_O32_H1, SVTN_682_954.time_O32_H2, SVTN_682_954.time_O32_R1, SVTN_682_954.time_O32_R2, SVTN_682_954.time_O32_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    
    # Controller2 installation
    # Pick, prepare and place controller2
    [SVTN_682_954.time_O41_H1, SVTN_682_954.time_O41_H2, SVTN_682_954.time_O41_R1, SVTN_682_954.time_O41_R2, SVTN_682_954.time_O41_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    # Tighten 4 screws
    [SVTN_682_954.time_O42_H1, SVTN_682_954.time_O42_H2, SVTN_682_954.time_O42_R1, SVTN_682_954.time_O42_R2, SVTN_682_954.time_O42_R3, 9999, 9999, 9999, 9999, 9999, 9999],

    # Cables installation
    [9999, 9999, 9999, 9999, 9999, SVTN_682_954.time_O51_H1, 9999, 9999, SVTN_682_954.time_O51_H2, 9999, 9999],
]

Time_H2R3_682_997 = [
    # Battery1 installation
    # Pick and place battery1
    [SVTN_682_997.time_O11_H1, SVTN_682_997.time_O11_H2, SVTN_682_997.time_O11_R1, SVTN_682_997.time_O11_R2, SVTN_682_997.time_O11_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    # Tighten 4 screws of left side
    [SVTN_682_997.time_O12_H1, SVTN_682_997.time_O12_H2, SVTN_682_997.time_O12_R1, SVTN_682_997.time_O12_R2, SVTN_682_997.time_O12_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    # Tighten 4 screws of right side
    [SVTN_682_997.time_O13_H1, SVTN_682_997.time_O13_H2, SVTN_682_997.time_O13_R1, SVTN_682_997.time_O13_R2, SVTN_682_997.time_O13_R3, 9999, 9999, 9999, 9999, 9999, 9999],

    # Battery2 installation
    # Pick and place battery2
    [SVTN_682_997.time_O21_H1, SVTN_682_997.time_O21_H2, SVTN_682_997.time_O21_R1, SVTN_682_997.time_O21_R2, SVTN_682_997.time_O21_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    # Tighten 4 screws of left side
    [SVTN_682_997.time_O22_H1, SVTN_682_997.time_O22_H2, SVTN_682_997.time_O22_R1, SVTN_682_997.time_O22_R2, SVTN_682_997.time_O22_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    # Tighten 4 screws of right side
    [SVTN_682_997.time_O23_H1, SVTN_682_997.time_O23_H2, SVTN_682_997.time_O23_R1, SVTN_682_997.time_O23_R2, SVTN_682_997.time_O23_R3, 9999, 9999, 9999, 9999, 9999, 9999],

    # Controller1 installation
    # Pick, prepare and place controller1
    [SVTN_682_997.time_O31_H1, SVTN_682_997.time_O31_H2, SVTN_682_997.time_O31_R1, SVTN_682_997.time_O31_R2, SVTN_682_997.time_O31_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    # Tighten 4 screws
    [SVTN_682_997.time_O32_H1, SVTN_682_997.time_O32_H2, SVTN_682_997.time_O32_R1, SVTN_682_997.time_O32_R2, SVTN_682_997.time_O32_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    
    # Controller2 installation
    # Pick, prepare and place controller2
    [SVTN_682_997.time_O41_H1, SVTN_682_997.time_O41_H2, SVTN_682_997.time_O41_R1, SVTN_682_997.time_O41_R2, SVTN_682_997.time_O41_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    # Tighten 4 screws
    [SVTN_682_997.time_O42_H1, SVTN_682_997.time_O42_H2, SVTN_682_997.time_O42_R1, SVTN_682_997.time_O42_R2, SVTN_682_997.time_O42_R3, 9999, 9999, 9999, 9999, 9999, 9999],

    # Cables installation
    [9999, 9999, 9999, 9999, 9999, SVTN_682_997.time_O51_H1, 9999, 9999, SVTN_682_997.time_O51_H2, 9999, 9999],
]

Time_H2R3_954_997 = [
    # Battery1 installation
    # Pick and place battery1
    [SVTN_954_997.time_O11_H1, SVTN_954_997.time_O11_H2, SVTN_954_997.time_O11_R1, SVTN_954_997.time_O11_R2, SVTN_954_997.time_O11_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    # Tighten 4 screws of left side
    [SVTN_954_997.time_O12_H1, SVTN_954_997.time_O12_H2, SVTN_954_997.time_O12_R1, SVTN_954_997.time_O12_R2, SVTN_954_997.time_O12_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    # Tighten 4 screws of right side
    [SVTN_954_997.time_O13_H1, SVTN_954_997.time_O13_H2, SVTN_954_997.time_O13_R1, SVTN_954_997.time_O13_R2, SVTN_954_997.time_O13_R3, 9999, 9999, 9999, 9999, 9999, 9999],

    # Battery2 installation
    # Pick and place battery2
    [SVTN_954_997.time_O21_H1, SVTN_954_997.time_O21_H2, SVTN_954_997.time_O21_R1, SVTN_954_997.time_O21_R2, SVTN_954_997.time_O21_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    # Tighten 4 screws of left side
    [SVTN_954_997.time_O22_H1, SVTN_954_997.time_O22_H2, SVTN_954_997.time_O22_R1, SVTN_954_997.time_O22_R2, SVTN_954_997.time_O22_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    # Tighten 4 screws of right side
    [SVTN_954_997.time_O23_H1, SVTN_954_997.time_O23_H2, SVTN_954_997.time_O23_R1, SVTN_954_997.time_O23_R2, SVTN_954_997.time_O23_R3, 9999, 9999, 9999, 9999, 9999, 9999],

    # Controller1 installation
    # Pick, prepare and place controller1
    [SVTN_954_997.time_O31_H1, SVTN_954_997.time_O31_H2, SVTN_954_997.time_O31_R1, SVTN_954_997.time_O31_R2, SVTN_954_997.time_O31_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    # Tighten 4 screws
    [SVTN_954_997.time_O32_H1, SVTN_954_997.time_O32_H2, SVTN_954_997.time_O32_R1, SVTN_954_997.time_O32_R2, SVTN_954_997.time_O32_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    
    # Controller2 installation
    # Pick, prepare and place controller2
    [SVTN_954_997.time_O41_H1, SVTN_954_997.time_O41_H2, SVTN_954_997.time_O41_R1, SVTN_954_997.time_O41_R2, SVTN_954_997.time_O41_R3, 9999, 9999, 9999, 9999, 9999, 9999],
    # Tighten 4 screws
    [SVTN_954_997.time_O42_H1, SVTN_954_997.time_O42_H2, SVTN_954_997.time_O42_R1, SVTN_954_997.time_O42_R2, SVTN_954_997.time_O42_R3, 9999, 9999, 9999, 9999, 9999, 9999],

    # Cables installation
    [9999, 9999, 9999, 9999, 9999, SVTN_954_997.time_O51_H1, 9999, 9999, SVTN_954_997.time_O51_H2, 9999, 9999],
]

def transTime_SVTNN2static(Time):
    new_Time = copy.deepcopy(Time)
    for lst in new_Time:
      for i in range(len(lst)):
          if lst[i] != 9999:
              lst[i] = lst[i].mean
    return new_Time
def transTime_SVTNN2fuzzy(Time):
    new_Time = copy.deepcopy(Time)
    for lst in new_Time:
      for i in range(len(lst)):
          if lst[i] != 9999:
            lst[i] = Triangular(lst[i].l_std, lst[i].mean, lst[i].r_std)
    return new_Time
def setWeight_SVTNN(Time, weight):
    new_Time = copy.deepcopy(Time)
    for lst in new_Time:
      for i in range(len(lst)):
          if lst[i] != 9999:
            lst[i].set_weight(weight, 1-weight, 1-weight)
    return new_Time

# station_static.set_operator_product_Time([human1, human2], [robot1, robot2, robot3], [product1, product2, product3], [transTime_SVTNN2static(Time_H2R3), transTime_SVTNN2static(Time_H2R3), transTime_SVTNN2static(Time_H2R3)])
# station_fuzzy.set_operator_product_Time([human1, human2], [robot1, robot2, robot3], [product4, product5, product6], [transTime_SVTNN2fuzzy(Time_H2R3), transTime_SVTNN2fuzzy(Time_H2R3), transTime_SVTNN2fuzzy(Time_H2R3)])

def excel_writing(filename, data):
    # 创建一个新的工作簿
    wb = Workbook()
    # 选择活动工作表
    ws = wb.active

    # 写入列标题
    column_titles = ['Iteration','Low', 'Optimal Fitness', 'Up','Low', 'Best Fitness', 'Up', 'Mean Fitness', 'Offspring Number', 'Time', 'Optimal Size']
    for col, title in enumerate(column_titles, start=1):
        ws.cell(row=1, column=col, value=title)

    # 写入具体数值
    for row, item in enumerate(data, start=2):
        for col, value in enumerate(item, start=1):
            ws.cell(row=row, column=col, value=value)

    # 保存工作簿
    if os.path.exists(filename):
        send2trash.send2trash(filename)  # 如果文件存在，先删除它
    wb.save(filename)

def printable(record_list):
    # record_list: [[iteration, low, optimal_fitness, up, low, best_fitness, up, mean_fitness, pop_num, time, optimal_size], ...]
    for record in record_list:
        optimal_fitness = record[1]
        if type(optimal_fitness) != SVTNN:
            best_fitness = record[2]
            record.insert(1, optimal_fitness)
            record.insert(3, optimal_fitness)
            record.insert(4, best_fitness)
            record.insert(6, best_fitness)
            record[2] = optimal_fitness
            record[5] = best_fitness
            for i in range(len(record)):
                record[i] = round(record[i])
        else:
            best_fitness = record[2]
            record.insert(1, optimal_fitness.low)
            record.insert(3, optimal_fitness.up)
            record.insert(4, best_fitness.low)
            record.insert(6, best_fitness.up)
            record[2] = optimal_fitness.mean
            record[5] = best_fitness.mean
            for i in range(len(record)):
                record[i] = round(record[i])

# station.gantt(OS_COS_MS_TS_START_END_AFTER)

# p_cross_machine 变异了MS就不会变异OS了
pop_size_list = [100, 200,300]
p_crossover_list = [0.7,0.8,0.9]
p_cross_machine_list = [0.5]
p_mutation_list = [0.1,0.2,0.3]
p_mutation_machine_list = [0.5]
max_iteration_list = [1000]
tSize_list = [3,4,5]



# 生成所有参数的组合
combinations = [
    [0.682, 0.954, 0.25],
    [0.682, 0.954, 0.50],
    [0.682, 0.954, 0.75],
    [0.682, 0.997, 0.25],
    [0.682, 0.997, 0.50],
    [0.682, 0.997, 0.75],
    [0.954, 0.997, 0.25],
    [0.954, 0.997, 0.50],
    [0.954, 0.997, 0.75],
]

path = "/home/amtc/Desktop/optimal/sensitivityTest/results/H2R3P3/" # 文件保存路径
mem_static_CHS_path = "/home/amtc/Desktop/optimal/sensitivityTest/results/H2R3P3/434-popSize=300_pCrossover=0.9_pCrossMS=0.5_pMutate=0.2_pMutateMS=0.5_maxIteration=1000.pickle"

# 遍历排列组合并生成相应的循环
for i in [1,2,3,4,5]:
    for alpha, beta, weight in combinations:

        product1 = Productt("Battery pack")
        product2 = Productt("Battery pack")
        product3 = Productt("Battery pack")
        product1.set_jobs(n_job_op, after_require)
        product2.set_jobs(n_job_op, after_require)
        product3.set_jobs(n_job_op, after_require)
        station_SVTNN = HrcStation()

        if alpha == 0.682 and beta == 0.954:
            Time = Time_H2R3_682_954
        elif alpha == 0.682 and beta == 0.997:
            Time = Time_H2R3_682_997
        elif alpha == 0.954 and beta == 0.997:
            Time = Time_H2R3_954_997
        Time = setWeight_SVTNN(Time, weight)
        station_SVTNN.set_operator_product_Time([human1, human2], [robot1, robot2, robot3], [product1, product2, product3], [Time, Time, Time])

        with open(mem_static_CHS_path, 'rb') as file: 
            # 使用pickle.load()函数加载数据
            static_CHS_list = pickle.load(file)
            mem_static_CHS_list = random.sample(static_CHS_list, 300)
        
        if __name__ == '__main__':
            optimal_fitness, mem_SVTNN_CHS_list, mem_SVTNN_CHS_fitness_list,  record_list = station_SVTNN.GA(pop_size=300, p_crossover=0.9, p_cross_machine=0.67, p_mutation=0.1, p_mutation_machine=0.83, max_iteration=1500, tSize=4)
            # optimal_fitness, mem_SVTNN_CHS_list, mem_SVTNN_CHS_fitness_list,  record_list = station_SVTNN.GA_initialed_pop(mem_static_CHS_list, pop_size=300, p_crossover=0.9, p_cross_machine=0.67, p_mutation=0.1, p_mutation_machine=0.83, max_iteration=300, tSize=4)
            
            printable(record_list)
            excel_name = path +"{}/".format(i) + "{}-alpha={}_beta={}_weight={}.xlsx".format(optimal_fitness.mean if isinstance(optimal_fitness, SVTNN) else optimal_fitness, alpha, beta, weight)
            data_name = path +"{}/".format(i) + "{}-alpha={}_beta={}_weight={}.pickle".format(optimal_fitness.mean if isinstance(optimal_fitness, SVTNN) else optimal_fitness, alpha, beta, weight)
            excel_writing(excel_name, record_list)
            with open(data_name, "wb") as file:
                pickle.dump(mem_SVTNN_CHS_list, file)